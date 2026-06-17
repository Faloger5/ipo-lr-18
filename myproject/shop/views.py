from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.contrib import messages
from django.core.mail import EmailMessage
from django.conf import settings
from django.core.paginator import Paginator
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime

from rest_framework import viewsets, permissions, generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated, AllowAny

from .models import Product, Category, Manufacturer, Cart, CartItem, Order, OrderItem, UserProfile
from .serializers import (
    CategorySerializer, ManufacturerSerializer,
    ProductSerializer, CartSerializer, CartItemSerializer,
    UserProfileSerializer, OrderSerializer
)
from .permissions import IsAdminOrManager, IsOwnerOrAdmin


def home(request):
    products = Product.objects.filter(stock__gt=0).order_by('-id')[:6]
    categories = Category.objects.all()
    return render(request, 'shop/index.html', {
        'products': products,
        'categories': categories,
    })


def about(request):
    """Страница об авторе"""
    return HttpResponse("""
        <h1>Об авторе</h1>
        <p><strong>Имя:</strong> Доброва Анна</p>
        <p><strong>Группа:</strong> 88ТП</p>
        <p><strong>Email:</strong> student@example.com</p>
        <p><a href="/">На главную</a></p>
    """)


def shop_info(request):
    """Страница о магазине"""
    return HttpResponse("""
        <h1>О магазине подарочных сертификатов</h1>
        <p><strong>Название:</strong> Surprise.by</p>
        <p><strong>Тема:</strong> Интернет-магазин подарочных сертификатов сетей Минска</p>
        <p><strong>Описание:</strong> Подарочные сертификаты в популярные сети магазинов Минска: 
        Евроопт, Грин, Соседи, Магнит, Belarusian Brand, Свитанок, 5 элемент.</p>
        <p><a href="/">На главную</a></p>
    """)


def product_list(request):
    products = Product.objects.all()

    search_query = request.GET.get('search', '')
    category_id = request.GET.get('category', '')
    manufacturer_id = request.GET.get('manufacturer', '')

    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    if category_id:
        products = products.filter(category_id=category_id)
    if manufacturer_id:
        products = products.filter(manufacturer_id=manufacturer_id)

    paginator = Paginator(products, 9)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'categories': Category.objects.all(),
        'manufacturers': Manufacturer.objects.all(),
        'search_query': search_query,
        'selected_category': category_id,
        'selected_manufacturer': manufacturer_id,
    }
    return render(request, 'shop/catalog.html', context)


def product_detail(request, pk):
    product = get_object_or_404(Product, id=pk)
    return render(request, 'shop/product_detail.html', {'product': product})


@login_required
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    cart, _ = Cart.objects.get_or_create(user=request.user)
    cart_item, created = CartItem.objects.get_or_create(
        cart=cart, product=product, defaults={'quantity': 1}
    )
    if not created:
        if cart_item.quantity + 1 <= product.stock:
            cart_item.quantity += 1
            cart_item.save()
            messages.success(request, f'Товар "{product.name}" добавлен в корзину')
        else:
            messages.error(request, f'Недостаточно товара. Доступно: {product.stock} шт.')
    else:
        messages.success(request, f'Товар "{product.name}" добавлен в корзину')
    return redirect('cart')


@login_required
def update_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    if request.method == 'POST':
        quantity = int(request.POST.get('quantity', 1))
        if quantity > cart_item.product.stock:
            messages.error(request, f'Недостаточно товара. Доступно: {cart_item.product.stock} шт.')
        elif quantity > 0:
            cart_item.quantity = quantity
            cart_item.save()
            messages.success(request, 'Количество обновлено')
        else:
            cart_item.delete()
            messages.success(request, 'Товар удалён из корзины')
    return redirect('cart')


@login_required
def remove_from_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    product_name = cart_item.product.name
    cart_item.delete()
    messages.success(request, f'Товар "{product_name}" удалён из корзины')
    return redirect('cart')


@login_required
def cart_view(request):
    cart, _ = Cart.objects.get_or_create(user=request.user)
    cart_items = cart.items.select_related('product').all()
    total_price = sum(item.item_price() for item in cart_items)
    return render(request, 'shop/cart.html', {
        'cart': cart,
        'cart_items': cart_items,
        'total_price': total_price,
    })


@login_required
def checkout(request):
    try:
        cart = request.user.cart
        cart_items = cart.items.select_related('product').all()
    except Cart.DoesNotExist:
        messages.error(request, 'Ваша корзина пуста')
        return redirect('cart')

    if not cart_items.exists():
        messages.error(request, 'Ваша корзина пуста')
        return redirect('cart')

    if request.method == 'POST':
        customer_email = request.POST.get('email', '').strip()
        address = request.POST.get('address', '').strip()
        phone = request.POST.get('phone', '').strip()
        comment = request.POST.get('comment', '').strip()

        errors = []
        if not customer_email:
            errors.append('Email обязателен')
        if not address:
            errors.append('Адрес обязателен')
        if not phone:
            errors.append('Телефон обязателен')
        if phone and not phone.startswith('+'):
            errors.append('Телефон должен начинаться с +')

        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'shop/checkout.html', {
                'cart_items': cart_items,
                'total_price': cart.total_price(),
                'email': customer_email,
                'address': address,
                'phone': phone,
                'comment': comment,
            })

        order = Order.objects.create(
            user=request.user,
            customer_email=customer_email,
            address=address,
            phone=phone,
            comment=comment,
            total_amount=cart.total_price(),
            status='pending',
        )

        for cart_item in cart_items:
            OrderItem.objects.create(
                order=order,
                product=cart_item.product,
                quantity=cart_item.quantity,
                price=cart_item.product.price,
            )
            cart_item.product.stock -= cart_item.quantity
            cart_item.product.save()

        excel_buffer = generate_receipt_excel(order)

        subject = f'Ваш заказ №{order.id} оформлен'
        body = (
            f'Здравствуйте, {request.user.username}!\n\n'
            f'Ваш заказ №{order.id} успешно оформлен.\n\n'
            f'Адрес доставки: {address}\n'
            f'Телефон: {phone}\n'
            f'Сумма заказа: {order.total_amount} BYN\n\n'
            f'Чек прикреплён к письму.\n\n'
            f'Спасибо за покупку!'
        )

        try:
            email = EmailMessage(
                subject=subject,
                body=body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[customer_email],
            )
            email.attach(
                filename=f'receipt_order_{order.id}.xlsx',
                content=excel_buffer.read(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            email.send(fail_silently=False)
            messages.success(request, f'Заказ №{order.id} оформлен! Чек отправлен на {customer_email}')
        except Exception as e:
            messages.warning(request, f'Заказ оформлен, но письмо не отправлено: {e}')

        cart_items.delete()
        return redirect('order_success', order_id=order.id)

    return render(request, 'shop/checkout.html', {
        'cart_items': cart_items,
        'total_price': cart.total_price(),
    })


def generate_receipt_excel(order):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Чек'

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    ws['A1'] = 'Surprise.by'
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')

    ws['A2'] = 'Подарочный сертификат'
    ws['A2'].font = header_font
    ws.merge_cells('A2:E2')

    ws['A3'] = f'Дата: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    ws.merge_cells('A3:E3')
    ws['A4'] = f'Заказ №: {order.id}'
    ws.merge_cells('A4:E4')

    ws['A6'] = 'Покупатель:'
    ws['A6'].font = header_font
    ws['B6'] = order.user.username
    ws['C6'] = 'Email:'
    ws['D6'] = order.customer_email

    ws['A7'] = 'Телефон:'
    ws['B7'] = order.phone
    ws['C7'] = 'Адрес:'
    ws['D7'] = order.address

    headers = ['№', 'Сертификат', 'Номинал', 'Кол-во', 'Сумма']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.font = header_font
        cell.border = border

    row = 10
    for idx, item in enumerate(order.items.all(), 1):
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=item.product.name).border = border
        ws.cell(row=row, column=3, value=float(item.price)).border = border
        ws.cell(row=row, column=4, value=item.quantity).border = border
        ws.cell(row=row, column=5, value=float(item.price * item.quantity)).border = border
        row += 1

    ws.cell(row=row, column=4, value='ИТОГО:').font = Font(bold=True)
    ws.cell(row=row, column=5, value=float(order.total_amount)).font = Font(bold=True)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def order_success(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    return render(request, 'shop/order_success.html', {'order': order})


@login_required
def profile_view(request):
    categories = Category.objects.all()
    return render(request, 'shop/profile.html', {'categories': categories})


# ───────────────────────── API ViewSets ─────────────────────────

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]


class ManufacturerViewSet(viewsets.ModelViewSet):
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            self.permission_classes = [IsAdminOrManager]
        else:
            self.permission_classes = [AllowAny]
        return super().get_permissions()

    def get_queryset(self):
        qs = Product.objects.all()
        params = self.request.query_params
        if params.get('category'):
            qs = qs.filter(category_id=params['category'])
        if params.get('manufacturer'):
            qs = qs.filter(manufacturer_id=params['manufacturer'])
        if params.get('search'):
            qs = qs.filter(name__icontains=params['search'])
        return qs


class CartViewSet(viewsets.ModelViewSet):
    serializer_class = CartSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Cart.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=True, methods=['post'])
    def add_item(self, request, pk=None):
        cart = self.get_object()
        product_id = request.data.get('product_id')
        quantity = int(request.data.get('quantity', 1))
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return Response({'error': 'Товар не найден'}, status=404)

        cart_item, created = CartItem.objects.get_or_create(
            cart=cart, product=product, defaults={'quantity': quantity}
        )
        if not created:
            if cart_item.quantity + quantity <= product.stock:
                cart_item.quantity += quantity
                cart_item.save()
            else:
                return Response({'error': 'Недостаточно товара на складе'}, status=400)

        return Response(CartItemSerializer(cart_item).data)

    @action(detail=True, methods=['delete'])
    def remove_item(self, request, pk=None):
        cart = self.get_object()
        try:
            CartItem.objects.get(id=request.data.get('item_id'), cart=cart).delete()
            return Response({'message': 'Товар удалён из корзины'})
        except CartItem.DoesNotExist:
            return Response({'error': 'Товар не найден в корзине'}, status=404)


class CartItemViewSet(viewsets.ModelViewSet):
    serializer_class = CartItemSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return CartItem.objects.filter(cart__user=self.request.user)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        quantity = request.data.get('quantity')
        if quantity and int(quantity) > instance.product.stock:
            return Response({'error': 'Недостаточно товара на складе'}, status=400)
        return super().update(request, *args, **kwargs)


class MeAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = UserProfileSerializer(request.user.profile)
        return Response(serializer.data)

    def patch(self, request):
        profile = request.user.profile
        serializer = UserProfileSerializer(profile, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class MyOrdersAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = OrderSerializer

    def get_queryset(self):
        if hasattr(self.request.user, 'profile') and self.request.user.profile.role == 'admin':
            return Order.objects.all()
        return Order.objects.filter(user=self.request.user)
    from django.shortcuts import render, redirect
from django.contrib.auth import login
from django.contrib import messages
from django.contrib.auth.forms import UserCreationForm
from django import forms
from django.contrib.auth.models import User
from django.views.decorators.csrf import ensure_csrf_cookie


class RegisterForm(UserCreationForm):
    email = forms.EmailField(required=False, label='Email')
    full_name = forms.CharField(max_length=200, required=False, label='Полное имя')
    phone = forms.CharField(max_length=20, required=False, label='Телефон')

    class Meta:
        model = User
        fields = ('username', 'email', 'password1', 'password2')

    def save(self, commit=True):
        user = super().save(commit=False)
        user.email = self.cleaned_data.get('email', '')
        if commit:
            user.save()
        return user


@ensure_csrf_cookie
def register(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            
            if not hasattr(user, 'profile'):
                from shop.models import UserProfile
                UserProfile.objects.create(user=user)
            
            profile = user.profile
            profile.full_name = form.cleaned_data.get('full_name', '')
            profile.phone = form.cleaned_data.get('phone', '')
            profile.save()

            login(request, user)
            messages.success(request, f'Добро пожаловать, {user.username}!')
            return redirect('catalog')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{error}')
    else:
        form = RegisterForm()

    return render(request, 'registration/register.html', {'form': form})